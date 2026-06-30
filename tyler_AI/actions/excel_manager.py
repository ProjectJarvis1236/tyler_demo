from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import httpx
import json
import win32com.client
import pandas as pd
from pathlib import Path

from .File_manager import FileDescriptionManager
import configs

API_KEY = configs.OPENROUTER_KEY
URL = configs.URLS["nvidia/nemotron-3-super-120b-a12b:free"]  # .env


class ExcelManager:
    async def run(self, params: dict, chat_id: str):
        print("зашёл")
        searcher = FileDescriptionManager(configs.USER_FOLDER)
        nm = params.get("filename")
        if not nm.endswith('.xlsx'): nm += '.xlsx'
        if params["action"] != "create":
            name = FileDescriptionManager.find_file_by_description(nm)
        else:
            name = nm

        folder = Path(configs.USER_FOLDER)
        filename = folder / name

        action = params.get("action")
        sheet_name = params.get("sheet", "Sheet")

        if action == "create":
            return self._create_file(filename, sheet_name, params.get("data"))
        if action == "update":
            return await self._update_file(filename, sheet_name, params.get("updates"))
        if action == "read":
            return await self._smart_parse(filename, sheet_name)
        return {"status": "error", "message": "Неизвестное действие в Excel"}

    def _create_file(self, filename, sheet_name, data):
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if data:
            headers = data.get("headers", [])
            rows = data.get("rows", [])
            if headers:
                ws.append(headers)
            for row in rows:
                ws.append(row)

        wb.save(str(filename))
        print(f"[DEBUG] Файл {filename} создан с листом {sheet_name}")
        return {"status": "ok", "message": f"Файл {filename} создан"}

    def get_table_bounds(self, ws):
        min_row, min_col = None, None
        max_row, max_col = 0, 0

        for row in ws.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    r, c = cell.row, cell.column

                    if min_row is None or r < min_row:
                        min_row = r
                    if min_col is None or c < min_col:
                        min_col = c

                    if r > max_row:
                        max_row = r
                    if c > max_col:
                        max_col = c

        return min_row, min_col, max_row, max_col

    async def _update_file(self, filename, sheet_name, updates):
        if not os.path.exists(filename):
            return {"status": "error", "message": "Файл не найден"}

        was_open = self._close_if_open(filename)

        mission = updates.get("mission")

        structure = await self._smart_parse(filename, sheet_name)

        wb = load_workbook(filename)
        ws = wb[sheet_name]

        start_row, start_col, end_row, end_col = self.get_table_bounds(ws)

        start_col_letter = get_column_letter(start_col)
        end_col_letter = get_column_letter(end_col)

        last_col_letter = get_column_letter(ws.max_column)
        safe_row_for_results = ws.max_row + 2

        prompt = {
            "model": "nvidia/nemotron-3-super-120b-a12b:free",
            "messages": [
                {
                    "role": "system",
                    "content": f"""
                    Ты эксперт по анализу данных в Excel.

                    Ты ОБЯЗАН выбрать правильный тип операции: add_column или aggregate.

                    Ты пишешь действия для работы с таблицей.

                    Формат ответа:
                    {{
                    "operations": [
                        {{
                        "type": "add_column",
                        "new_column": "Название",
                        "excel_formula": "A2+B2"
                        }},
                        {{
                        "type": "aggregate",
                        "excel_formula": "SUM(A2:A1000)",
                        "target_cell": "D1002"
                        }}
                    ]
                    }}
                    Используй add_column ТОЛЬКО если:
                    - нужно вычисление для КАЖДОЙ строки
                    - результат должен быть в новом столбце
                    - формула зависит от значений строки (A2, B2 и т.д.)

                    Используй aggregate ТОЛЬКО если:
                    - нужен ОДИН результат (одно число или текст)
                    - это итог, сумма, среднее, максимум и т.д.
                    - результат НЕ должен повторяться в каждой строке

                    При использовании aggregate обязательно пиши в клеточке пояснение формулы, а ниже саму формулу
                    Обязательно сначала определи: нужен один результат или значения для каждой строки
                    Только после этого выбирай тип операции.

                    Границы таблицы:
                    - Верхний левый угол: {start_col_letter}{start_row}
                    - Нижний правый угол: {end_col_letter}{end_row}

                    КРИТИЧЕСКО:
                    - Все диапазоны должны заканчиваться на строке {end_row}
                    - НЕ использовать строки ниже {end_row}

                    Правила:
                    - Используй в форрмулах ТОЛЬКО адреса ячеек (A2, B2)
                    - НЕ используй в формулах названия колонок
                    - НЕ используй @
                    - Формулы должны быть валидными Excel
                    - Пиши понятные названия колонок, которые объясняют суть столбца
                    - Заранее рассчитывай, какие колонки ты создашь, чтобы не писать формулы в эти колонки и не перекрывать другие значения
                    - В клеточки через операцию aggregate можешь писать текстовые пояснения к формулам, по необходимости
                    """
                },
                {
                    "role": "user",
                    "content": f"""
                        Задача:
                        {mission}

                        Структура таблицы по столбцам:
                        {json.dumps(structure['columns'], ensure_ascii=False)}

                        Ограничения:
                        - Новые колонки создавай после {last_col_letter}
                        """
                }
            ]
        }

        headers = {
            "Authorization": f"Bearer {API_KEY}",
            "Content-Type": "application/json"
        }

        print(f"[DEBUG] Отправляю запрос в LLM для файла {filename}...")
        print(f"[DEBUG] Mission: {mission}")
        # asyncio.sleep(0.5) #задержка, чтобы промпты успели создаться
        async with httpx.AsyncClient(timeout=configs.TIMEOUT) as client:
            response = await client.post(URL, headers=headers, json=prompt)
            if response.status_code != 200:
                return {"status": "error", "message": f"HTTP error {response.status_code}: {response.text}"}

        print("RAW:", response.text)

        result = response.json()
        raw = result.get("choices", [{}])[0].get("message", {}).get("content", "")
        raw = raw.replace("```json", "").replace("```", "").strip()

        try:
            parsed = json.loads(raw)
            operations = parsed.get("operations", [])
        except Exception as e:
            return {"status": "error", "message": f"Ошибка LLM: {e}"}

        if not operations:
            return {"status": "error", "message": "Нет операций"}

        result = self._execute_operation(filename, sheet_name, operations)

        if was_open: self._reopen_file(filename)

        return result

    def _execute_operation(self, filename, sheet_name, operations):
        wb = load_workbook(filename)
        ws = wb[sheet_name]

        def clean_formula(f):
            f = f.replace("@", "")
            f = f.replace("#ССЫЛКА!", "")
            f = f.strip()
            if not f.startswith("="): f = "=" + f
            return f

        initial_max_col = len([cell for cell in ws[1] if cell.value is not None])
        current_new_col = initial_max_col + 2

        for operation in operations:
            op_type = operation.get("type")

            if op_type == "add_column":
                new_column = operation["new_column"]
                formula_template = operation["excel_formula"]

                cleaned_formula = clean_formula(formula_template)
                new_col_letter = get_column_letter(current_new_col)
                new_header_addr = f"{new_col_letter}1"

                ws[new_header_addr].value = new_column
                print(f"[DEBUG] Добавляю колонку '{new_column}' в {new_header_addr}")

                for row_num in range(2, ws.max_row + 1):
                    formula = cleaned_formula

                    import re
                    def replace_cell(match):
                        col_letter = match.group(1)
                        return f"{col_letter}{row_num}"

                    pattern = r'([A-Z]+)(\d+)'
                    updated_formula = re.sub(pattern, replace_cell, formula)

                    cell_addr = f"{new_col_letter}{row_num}"
                    if not updated_formula.startswith('='):
                        updated_formula = f"={updated_formula}"

                    # Оборачивание формулы в IFERROR
                    wrapped_formula = f'=IFERROR({updated_formula[1:]}, "")'

                    ws[cell_addr].value = wrapped_formula
                    print(f"[DEBUG] {cell_addr}: formula -> {wrapped_formula}")

                # Очистка ошибок
                print(f"[CLEANUP] Проверяю столбец {new_col_letter} на наличие ошибок...")
                errors_found = 0
                wb.save(filename)

                wb_data = load_workbook(filename, data_only=True)
                ws_data = wb_data[sheet_name]

                for row_num in range(2, ws_data.max_row + 1):
                    cell_addr = f"{new_col_letter}{row_num}"
                    cell = ws_data[cell_addr]
                    value = cell.value
                    if isinstance(value, str) and value.startswith('#') and value.endswith('!'):
                        print(f"[CLEANUP] Найдена ошибка в {cell_addr}: {value}")
                        ws[cell_addr].value = None
                        errors_found += 1

                print(f"[CLEANUP] Удалено {errors_found} ошибочных значений в столбце {new_col_letter}.")

                current_new_col += 1

            elif op_type == "aggregate":
                formula_template = operation["excel_formula"]
                target_cell = operation["target_cell"]

                if not formula_template.startswith('='):
                    formula_template = f"={formula_template}"
                ws[target_cell].value = formula_template
                print(f"[DEBUG] Aggregate в {target_cell}: {formula_template}")

            elif op_type == "set_formula":
                cell_addr = operation["cell"]
                formula = operation["formula"]
                if not formula.startswith('='):
                    formula = f"={formula}"
                ws[cell_addr].value = formula
                print(f"[DEBUG] Set formula в {cell_addr}: {formula}")

            else:
                print(f"[DEBUG] Пропущена неизвестная операция: {op_type}")

        wb.save(filename)
        print(f"[DEBUG] Файл {filename} сохранён после выполнения операций")
        return {"status": "ok", "message": f"Выполнено {len(operations)} операций в файле {filename}"}

    async def _smart_parse(self, filename, sheet_name):
        df = pd.read_excel(filename, sheet_name=sheet_name)
        columns = []

        for i, col in enumerate(df.columns):
            series = df[col]

            col_info = {
                "name": str(col),
                "letter": get_column_letter(i + 1),
                "non_null": int(series.count())
            }

            if pd.api.types.is_numeric_dtype(series):
                col_info["type"] = "number"
                col_info["mean"] = float(series.mean()) if series.count() > 0 else None
                col_info["min"] = float(series.min()) if series.count() > 0 else None
                col_info["max"] = float(series.max()) if series.count() > 0 else None
            else:
                col_info["type"] = "string"
                col_info["unique"] = int(series.nunique())

            columns.append(col_info)

        return {
            "status": "ok",
            "columns": columns,
            "rows": len(df),
        }

    def _close_if_open(self, filename):
        try:
            excel = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            return False

        filename = os.path.abspath(filename)

        for wb in excel.Workbooks:
            if os.path.abspath(wb.FullName) == filename:
                print(f"[DEBUG] Закрываю открытый файл: {filename}")
                wb.Close(SaveChanges=True)
                return True
        return False

    def _reopen_file(self, filename):
        try:
            excel = win32com.client.Dispatch("Excel.Application")
            excel.Visible = True
            excel.Workbooks.Open(os.path.abspath(filename))
            print(f"[DEBUG] Закрываю открытый файл: {filename}")
        except Exception as e:
            print(f"[DEBUG] Файл заново открыт: {filename}")

    @staticmethod
    def _find_next_empty_cell(ws, start_cell):
        col_letter = ''.join(filter(str.isalpha, start_cell))
        row = int(''.join(filter(str.isdigit, start_cell)))
        while ws[f"{col_letter}{row}"].value is not None:
            row += 1
        return f"{col_letter}{row}"
